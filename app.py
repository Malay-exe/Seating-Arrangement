import matplotlib
# Set the backend to Agg for non-GUI environments
matplotlib.use('Agg')

import openpyxl
import matplotlib.pyplot as plt
import numpy as np
import io
import base64
from flask import Flask, render_template, request

app = Flask(__name__)

workbook = openpyxl.load_workbook('C:\Python Programs\Milu\ok.xlsx')
sheet = workbook['Sheet1'] 

def extract_seating_data():
    seating_data = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_data = []
        for cell in row:
            if cell.value:
                row_data.append(cell.value)
            else:
                row_data.append(None)  
        seating_data.append(row_data)
    return np.array(seating_data, dtype=object)

def generate_table_with_highlight(roll_number):
    seating_data_array = extract_seating_data()

    colors = [['#f5f5f5' for _ in range(seating_data_array.shape[1])] for _ in range(seating_data_array.shape[0])]  # Light gray background

    highlight_color = '#ffcc00' 

    for i, row in enumerate(seating_data_array):
        for j, cell in enumerate(row):
            if cell and str(roll_number) in str(cell):  
                colors[i][j] = highlight_color  

    fig, ax = plt.subplots(figsize=(14, 8)) 

    ax.axis('tight')
    ax.axis('off')

    table = ax.table(
        cellText=seating_data_array, 
        loc='center', 
        cellLoc='center', 
        colLabels=[f"Seat {i+1}" for i in range(seating_data_array.shape[1])], 
        rowLabels=[f"Row {i+1}" for i in range(seating_data_array.shape[0])],  
        cellColours=colors,
        colWidths=[0.12] * seating_data_array.shape[1]  
    )

    for (i, j), cell in table.get_celld().items():
        if i == 0:
            cell.set_fontsize(12)
            cell.set_text_props(weight='bold', color='white')
            cell.set_facecolor('#4CAF50') 
        else:
            cell.set_fontsize(10)
            cell.set_text_props(weight='normal', color='black') 
            if j == 0:
                cell.set_text_props(weight='bold')

            cell.set_edgecolor('black')  
            cell.set_linewidth(1) 

    img_io = io.BytesIO()
    plt.savefig(img_io, format='png', dpi=300)  
    img_io.seek(0)

    img_base64 = base64.b64encode(img_io.getvalue()).decode('utf-8')
    plt.close(fig)  

    return img_base64


@app.route('/', methods=['GET', 'POST'])
def front():
    return render_template('front.html')

@app.route('/index', methods=['GET', 'POST'])
def index():
    img_base64 = None
    if request.method == 'POST':
        roll_number = request.form.get('roll_number')
        if roll_number:
            img_base64 = generate_table_with_highlight(roll_number)
    return render_template('index.html', img_base64=img_base64)

# Run the Flask app
if __name__ == '__main__':
    app.run(debug=True)
